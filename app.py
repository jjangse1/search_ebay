"""
eBay 가격 조회 도구 - 로컬 서버
================================
myhubon.com 상품 목록(xlsx)을 업로드하면 가격이 등록되지 않은(price_display_yn != 'Y'
또는 가격 <= 1) 항목만 추려서 보여주고, 사용자가 선택한 항목에 대해서만 실제 eBay
Browse API(item_summary/search)를 호출해 현재 시세(최저/최고/중앙값)를 가져온다.

실행 방법 (Python으로 직접 실행)
---------------------------------
1) pip install -r requirements.txt
2) .env 파일 생성 (.env.example 참고) 후 EBAY_CLIENT_ID / EBAY_CLIENT_SECRET 입력
   - eBay Developer Program(https://developer.ebay.com) 가입 -> Application Keys 발급
   - Production Keyset 사용 시 실제 판매 중인 매물 시세 조회 가능
   - Sandbox Keyset은 테스트용 더미 데이터만 반환하므로 실거래 시세용으로는 사용 불가
3) python app.py
4) 브라우저에서 http://localhost:5050 접속 (자동으로도 열림)

exe로 빌드해서 쓰는 방법은 build.bat / README.md 참고.

이 서버는 로컬(localhost)에서만 동작하며, eBay Client Secret은 서버 프로세스 안에만
존재하고 브라우저로 절대 전송되지 않는다 (브라우저에서 직접 eBay Identity API를 호출하면
CORS로 막히고, Client Secret이 그대로 노출되는 문제가 있어 반드시 서버를 거쳐야 한다).
"""
import os
import re
import sys
import io
import gc
import time
import base64
import tempfile
import sqlite3
import statistics
import threading
import webbrowser
from urllib.parse import quote
from datetime import datetime, timedelta

import requests
from flask import Flask, request, jsonify, send_from_directory, session, redirect
from openpyxl import load_workbook
from dotenv import load_dotenv


def app_base_dir():
    """실행 방식(python 스크립트 vs PyInstaller onefile exe)에 따라
    .env 파일을 찾을 기준 폴더를 결정한다 (exe 파일이 있는 폴더)."""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def resource_path(rel_path):
    """static/ 같은 번들 리소스 경로. PyInstaller onefile은 실행 시 임시 폴더
    (sys._MEIPASS)에 리소스를 풀어놓으므로 그 경로를 우선 사용한다."""
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, rel_path)


BASE_DIR = app_base_dir()
load_dotenv(os.path.join(BASE_DIR, ".env"))

# ---------------------------------------------------------------------------
# 가격 이력 저장용 SQLite DB. exe/스크립트가 있는 폴더에 파일로 저장되어
# 서버를 재시작해도 유지되고, 이 서버에 접속하는 모든 사람이 같은 이력을 본다
# (기존 브라우저 localStorage 방식과 달리 팀 전체 공유).
#
# 주의(Render 등 클라우드 배포 시): Render 기본 디스크는 배포/재시작마다
# 초기화되는 "ephemeral(휘발성)" 파일시스템이라, DB_PATH 환경변수 없이 코드 위치
# 옆에 그냥 저장하면 재배포할 때마다 가격 추이 기록이 날아간다. Render에서
# Persistent Disk를 추가하고(예: /var/data 마운트) 환경변수 DB_PATH=/var/data/price_history.db
# 로 지정하면 재배포해도 기록이 유지된다.
# ---------------------------------------------------------------------------
DB_PATH = os.getenv("DB_PATH") or os.path.join(BASE_DIR, "price_history.db")


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS price_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            part_no TEXT NOT NULL,
            brand TEXT,
            query TEXT,
            count INTEGER,
            min_usd REAL,
            max_usd REAL,
            median_usd REAL,
            marketplaces TEXT,
            created_at TEXT NOT NULL
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_price_history_part_no ON price_history(part_no)")
    conn.commit()
    conn.close()


init_db()


def record_price_history(part_no, brand, query, stat, marketplaces):
    """검색 결과에 usd_stats가 있을 때마다 자동으로 한 줄 기록. 조회 자체를
    막지 않도록 저장 중 오류가 나도 조용히 무시한다."""
    if not part_no or not stat:
        return
    try:
        conn = get_db()
        conn.execute(
            "INSERT INTO price_history (part_no, brand, query, count, min_usd, max_usd, median_usd, marketplaces, created_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (
                str(part_no).strip().upper(),
                brand or "",
                query or "",
                stat["count"],
                stat["min"],
                stat["max"],
                stat["median"],
                ",".join(marketplaces or []),
                datetime.utcnow().isoformat() + "Z",
            ),
        )
        conn.commit()
        conn.close()
    except Exception:
        pass

EBAY_CLIENT_ID = os.getenv("EBAY_CLIENT_ID", "")
EBAY_CLIENT_SECRET = os.getenv("EBAY_CLIENT_SECRET", "")
EBAY_ENV = os.getenv("EBAY_ENV", "production").lower()   # "production" or "sandbox"

BASE_URL = "https://api.sandbox.ebay.com" if EBAY_ENV == "sandbox" else "https://api.ebay.com"

app = Flask(__name__, static_folder=resource_path("static"), static_url_path="")
# 업로드 파일 크기 상한 - Render 같은 메모리 적은(512MB) 서버에서 지나치게 큰 파일이
# 올라와서 그대로 OOM(메모리 부족)으로 죽는 걸 방지. 상품 목록 엑셀 용도로는 30MB면 충분.
app.config["MAX_CONTENT_LENGTH"] = 30 * 1024 * 1024



# ---------------------------------------------------------------------------
# 로그인 - APP_USERNAME/APP_PASSWORD를 설정하면 예쁜 로그인 페이지(static/login.html)를
# 거쳐야 나머지 기능을 쓸 수 있다. 브라우저 기본 인증창(Basic Auth) 대신 세션 쿠키 방식.
# 공개 URL로 배포할 때(Render 등) 반드시 설정 권장 (안 그러면 누구나 내 eBay 키로
# API를 대신 호출할 수 있음).
# ---------------------------------------------------------------------------
APP_USERNAME = os.getenv("APP_USERNAME", "")
APP_PASSWORD = os.getenv("APP_PASSWORD", "")

# 세션 서명에 쓰는 키. 반드시 환경변수로 고정값을 넣어주는 게 좋다 (안 넣으면 서버
# 재시작/재배포마다 로그인이 풀림). render.yaml에서 자동 생성되도록 설정해둠.
app.secret_key = os.getenv("SECRET_KEY") or os.urandom(32)
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=30)

# 로그인 없이 접근 가능한 경로 (로그인 페이지 자체와 로그인 처리 API)
PUBLIC_PATHS = {"/login", "/api/login"}


@app.before_request
def require_login():
    if not APP_USERNAME or not APP_PASSWORD:
        return  # 로컬 개발 등 미설정 시 로그인 생략
    if request.path in PUBLIC_PATHS:
        return
    if session.get("authed"):
        return
    # API 호출은 리다이렉트 대신 401 JSON (프론트 fetch가 그대로 에러 처리하게)
    if request.path.startswith("/api/"):
        return jsonify({"error": "로그인이 필요합니다"}), 401
    return redirect(f"/login?next={quote(request.full_path.rstrip('?'))}")


@app.route("/login", methods=["GET"])
def login_page():
    return send_from_directory(resource_path("static"), "login.html")


@app.route("/api/login", methods=["POST"])
def api_login():
    body = request.get_json(force=True, silent=True) or {}
    username = (body.get("username") or "").strip()
    password = body.get("password") or ""
    if username == APP_USERNAME and password == APP_PASSWORD:
        session["authed"] = True
        session.permanent = True
        return jsonify({"ok": True})
    return jsonify({"error": "아이디 또는 비밀번호가 올바르지 않습니다"}), 401


@app.route("/api/logout", methods=["POST"])
def api_logout():
    session.clear()
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# eBay OAuth (Client Credentials Grant) - Application access token, 2시간 캐시
# ---------------------------------------------------------------------------
_token_cache = {"access_token": None, "expires_at": datetime.min}


def get_ebay_token():
    if not EBAY_CLIENT_ID or not EBAY_CLIENT_SECRET:
        raise RuntimeError(
            "EBAY_CLIENT_ID / EBAY_CLIENT_SECRET이 설정되지 않았습니다. .env 파일을 확인하세요."
        )
    now = datetime.utcnow()
    if _token_cache["access_token"] and now < _token_cache["expires_at"]:
        return _token_cache["access_token"]

    creds = base64.b64encode(f"{EBAY_CLIENT_ID}:{EBAY_CLIENT_SECRET}".encode()).decode()
    resp = requests.post(
        f"{BASE_URL}/identity/v1/oauth2/token",
        headers={
            "Content-Type": "application/x-www-form-urlencoded",
            "Authorization": f"Basic {creds}",
        },
        data={
            "grant_type": "client_credentials",
            "scope": "https://api.ebay.com/oauth/api_scope",
        },
        timeout=15,
    )
    resp.raise_for_status()
    data = resp.json()
    _token_cache["access_token"] = data["access_token"]
    # 여유 있게 60초 일찍 만료 처리
    _token_cache["expires_at"] = now + timedelta(seconds=data.get("expires_in", 7200) - 60)
    return _token_cache["access_token"]


# ---------------------------------------------------------------------------
# 조회 가능한 eBay 국가별 사이트(마켓플레이스) 목록
# ---------------------------------------------------------------------------
MARKETPLACES = {
    # "월드와이드" 가상 옵션 - 국가 필터를 아예 안 걸고 EBAY_US 사이트를 조회.
    # eBay Browse API는 itemLocationCountry 필터를 안 주면 KR로 배송 가능한
    # "전세계 모든 출고국" 매물을 다 돌려준다. 즉 이게 곧 월드와이드 검색이다.
    "EBAY_WORLDWIDE": {"label": "🌍 월드와이드 (전체 출고국, ebay.com)", "currency": "USD"},
    "EBAY_US": {"label": "미국 (ebay.com)",       "currency": "USD"},
    # 주요 국가 필터 옵션 (EBAY_US 기반) - 월드와이드 결과 중 특정 출고국만 보고 싶을 때
    "EBAY_US_JP": {"label": "일본 출고 (ebay.com - JP)", "currency": "USD"},
    "EBAY_US_TW": {"label": "대만 출고 (ebay.com - TW)", "currency": "USD"},
    "EBAY_US_CN": {"label": "중국 출고 (ebay.com - CN)", "currency": "USD"},
    "EBAY_US_KR": {"label": "한국 출고 (ebay.com - KR)", "currency": "USD"},
    # 기존 사이트별 마켓플레이스 (다른 나라 로컬 eBay 사이트)
    "EBAY_GB": {"label": "영국 (ebay.co.uk)",      "currency": "GBP"},
    "EBAY_DE": {"label": "독일 (ebay.de)",         "currency": "EUR"},
    "EBAY_FR": {"label": "프랑스 (ebay.fr)",       "currency": "EUR"},
    "EBAY_IT": {"label": "이탈리아 (ebay.it)",     "currency": "EUR"},
    "EBAY_ES": {"label": "스페인 (ebay.es)",       "currency": "EUR"},
    "EBAY_AU": {"label": "호주 (ebay.com.au)",     "currency": "AUD"},
    "EBAY_CA": {"label": "캐나다 (ebay.ca)",       "currency": "CAD"},
    "EBAY_HK": {"label": "홍콩 (ebay.com.hk)",     "currency": "HKD"},
    "EBAY_SG": {"label": "싱가포르 (ebay.com.sg)", "currency": "SGD"},
}
# "월드와이드" 값들이 겹치므로(EBAY_US에 국가필터 없음 = 월드와이드) 기본값은 월드와이드 하나만.
DEFAULT_MARKETPLACES = [m.strip() for m in os.getenv("EBAY_MARKETPLACES", os.getenv("EBAY_MARKETPLACE", "EBAY_WORLDWIDE")).split(",") if m.strip()]


@app.route("/api/marketplaces", methods=["GET"])
def list_marketplaces():
    return jsonify({"marketplaces": MARKETPLACES, "default": DEFAULT_MARKETPLACES})


# ---------------------------------------------------------------------------
# 실시간 환율 (ECB 기반, 무료/키불필요 Frankfurter API) - USD 환산용, 6시간 캐시
# ---------------------------------------------------------------------------
_fx_cache = {"rates_to_usd": {"USD": 1.0}, "fetched_at": datetime.min}


def get_usd_rates():
    now = datetime.utcnow()
    if _fx_cache["rates_to_usd"] and now - _fx_cache["fetched_at"] < timedelta(hours=6):
        return _fx_cache["rates_to_usd"]
    try:
        currencies = ",".join(sorted({m["currency"] for m in MARKETPLACES.values()} - {"USD"} | {"KRW"}))
        resp = requests.get(
            "https://api.frankfurter.app/latest",
            params={"from": "USD", "to": currencies},
            timeout=10,
        )
        resp.raise_for_status()
        data = resp.json()
        rates_from_usd = data.get("rates", {})  # 1 USD = X <currency>
        rates_to_usd = {"USD": 1.0}
        for cur, rate in rates_from_usd.items():
            if rate:
                rates_to_usd[cur] = 1.0 / rate  # 1 <currency> = ? USD
        _fx_cache["rates_to_usd"] = rates_to_usd
        _fx_cache["fetched_at"] = now
    except Exception:
        # 환율 조회 실패 시 캐시된 값(또는 USD만)이라도 반환 - 조회 자체가 막히지 않게
        pass
    return _fx_cache["rates_to_usd"]


# ---------------------------------------------------------------------------
# 1) 엑셀 업로드 -> 가격 없는(price_display_yn != 'Y' 또는 price<=1) 행 추출
# ---------------------------------------------------------------------------
REQUIRED_COLS = [
    "product_title", "product_price", "price_display_yn", "conditions",
    "part_no", "brand",
]


@app.route("/api/parse-excel", methods=["POST"])
def parse_excel():
    if "file" not in request.files:
        return jsonify({"error": "파일이 없습니다"}), 400
    f = request.files["file"]

    # ------------------------------------------------------------------
    # 메모리 절약 포인트 1: f.read()로 파일 전체를 메모리(BytesIO)에 올리는 대신
    # 디스크의 임시 파일에 1MB씩 나눠 써서 저장한다. 업로드 스트림 전체를
    # 한 번에 메모리로 복사하지 않아도 되므로 큰 파일에서 메모리 사용량이 확 줄어든다.
    # ------------------------------------------------------------------
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    no_price_rows = []
    priced_count = 0
    try:
        with os.fdopen(tmp_fd, "wb") as tmp:
            while True:
                chunk = f.stream.read(1024 * 1024)
                if not chunk:
                    break
                tmp.write(chunk)

        # ------------------------------------------------------------------
        # 메모리 절약 포인트 2: read_only=True로 열면 openpyxl이 시트 전체를
        # 셀 객체(서식/스타일 포함)로 메모리에 다 올리지 않고, 행 단위로 필요한
        # 만큼만 스트리밍해서 읽는다. 기본 모드(read_only=False)는 엑셀이 조금만
        # 커도 메모리를 몇 배로 잡아먹어서 Render 같은 저메모리 환경에서 OOM이
        # 나는 주된 원인이었다.
        # ------------------------------------------------------------------
        wb = load_workbook(tmp_path, read_only=True, data_only=True)
        try:
            ws = wb.active

            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if header_row is None:
                return jsonify({"error": "빈 엑셀 파일입니다"}), 400
            header = list(header_row)
            missing = [c for c in REQUIRED_COLS if c not in header]
            if missing:
                return jsonify({"error": f"필수 컬럼이 없습니다: {missing}"}), 400
            idx = {name: header.index(name) for name in header if name is not None}

            row_id = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_id += 1
                part_no = row[idx["part_no"]]
                if not part_no:
                    continue
                try:
                    price = float(row[idx["product_price"]])
                except (TypeError, ValueError):
                    price = None
                disp = row[idx["price_display_yn"]]
                if disp == "Y" and price is not None and price > 1:
                    # 가격 등록된 행은 목록에 안 쓰므로 개수만 세고 내용은 버린다
                    # (예전엔 이 행들도 다 리스트에 쌓아뒀다가 개수만 리턴해서 낭비였음)
                    priced_count += 1
                    continue
                no_price_rows.append({
                    "row_id": row_id,
                    "part_no": str(part_no),
                    "brand": row[idx["brand"]] or "",
                    "conditions": row[idx["conditions"]] or "",
                    "product_title": row[idx["product_title"]] or "",
                    "product_price": price,
                    "price_display_yn": disp,
                })
        finally:
            wb.close()
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        gc.collect()

    return jsonify({
        "no_price_count": len(no_price_rows),
        "priced_count": priced_count,
        "no_price_rows": no_price_rows,
    })


# ---------------------------------------------------------------------------
# 2) 선택된 항목 -> 실제 eBay Browse API 검색 -> 시세 요약
# ---------------------------------------------------------------------------
def trimmed_median(prices, min_count_to_trim=5):
    """매물이 충분히 많을 때(기본 5건 이상)만 최저 1건·최고 1건을 제외하고
    나머지로 중앙값을 계산한다. 표본이 적으면(4건 이하) 트리밍하면 정보 손실이
    너무 커지므로 원본 전체로 중앙값을 구한다.
    반환값: (trimmed_median, was_trimmed: bool)
    """
    n = len(prices)
    if n >= min_count_to_trim:
        s = sorted(prices)
        trimmed = s[1:-1]  # 최저 1개, 최고 1개 제외
        return statistics.median(trimmed), True
    return statistics.median(prices), False


def search_ebay_once(query, marketplace_id, limit=20, item_location_country=None):
    token = get_ebay_token()
    headers = {
        "Authorization": f"Bearer {token}",
        "X-EBAY-C-MARKETPLACE-ID": marketplace_id,
        # 배송지 국가를 한국(KR)으로 지정하여 한국으로 배송 가능한 모든 해외 셀러 매물 수집
        "X-EBAY-C-ENDUSERCTX": "contextualLocation=country=KR", 
    }
    
    params = {
        "q": query, 
        "limit": limit, 
        "sort": "price"
    }
    
    # 특정 출고국가(예: JP) 지정이 요청되었을 경우 필터 추가
    if item_location_country:
        params["filter"] = f"itemLocationCountry:{item_location_country}"

    resp = requests.get(
        f"{BASE_URL}/buy/browse/v1/item_summary/search",
        headers=headers, params=params, timeout=20,
    )
    if resp.status_code == 401:
        # 토큰 만료 등 -> 한 번 강제 재발급 후 재시도
        _token_cache["access_token"] = None
        headers["Authorization"] = f"Bearer {get_ebay_token()}"
        resp = requests.get(
            f"{BASE_URL}/buy/browse/v1/item_summary/search",
            headers=headers, params=params, timeout=20,
        )
    resp.raise_for_status()
    return resp.json()

def dedupe_listings(listings):
    """여러 국가(마켓플레이스) 사이트로 나눠서 조회하다 보면, 같은 판매자가 올린
    완전히 동일한 매물(itemId 동일)이 국가마다 한 번씩 중복으로 잡히는 경우가 있다
    (특히 유럽/월드와이드처럼 같은 글로벌 리스팅이 여러 로컬 사이트에 다 노출되는 경우).
    itemId 기준으로 첫 등장한 것만 대표로 남기고, 나머지는 "이 매물이 또 뜬 나라 목록"
    (duplicate_marketplaces)으로만 접어서 대표 항목에 붙여둔다 - 목록/통계는 깔끔해지고,
    어느 국가들에 중복으로 올라와 있었는지는 CSV에서 그대로 확인할 수 있다.
    """
    seen = {}
    order = []
    for l in listings:
        key = l.get("item_id") or l["itemWebUrl"]
        if key not in seen:
            l["duplicate_marketplaces"] = []
            seen[key] = l
            order.append(key)
        else:
            first = seen[key]
            label = l.get("marketplace_label")
            if label and label not in first["duplicate_marketplaces"] and label != first.get("marketplace_label"):
                first["duplicate_marketplaces"].append(label)
    return [seen[k] for k in order]


def extract_item_id(it_summary):
    """중복 판정에 쓸 매물 고유 ID. eBay Browse API의 itemId는 마켓플레이스(국가)가
    달라도 같은 실제 리스팅이면 동일한 값을 준다. 혹시 없으면 itemWebUrl에서 상품번호를
    뽑아 대신 쓴다."""
    item_id = it_summary.get("itemId")
    if item_id:
        return item_id
    url = it_summary.get("itemWebUrl") or ""
    m = re.search(r"/itm/(\d+)", url)
    return m.group(1) if m else None


@app.route("/api/search-ebay", methods=["POST"])
def search_ebay():
    body = request.get_json(force=True)
    items = body.get("items", [])  # [{row_id, query}, ...]
    marketplaces = body.get("marketplaces") or DEFAULT_MARKETPLACES
    marketplaces = [m for m in marketplaces if m in MARKETPLACES] or ["EBAY_US"]
    if not items:
        return jsonify({"error": "선택된 항목이 없습니다"}), 400

    fx = get_usd_rates()  # {currency: 1 currency = ? USD}

    results = []
    for it in items:
        query = (it.get("query") or "").strip()
        row_id = it.get("row_id")
        if not query:
            results.append({"row_id": row_id, "query": query, "error": "검색어가 비어 있습니다"})
            continue

        listings = []
        total_matching = 0
        warnings = []
        had_success = False

        for mkt in marketplaces:
            try:
                # EBAY_US_XX 형태인 경우 EBAY_US를 마켓으로 쓰되, XX를 국가 필터로 적용
                if mkt == "EBAY_WORLDWIDE":
                    # 월드와이드 = EBAY_US 사이트를 국가 필터 없이 그대로 조회
                    actual_mkt = "EBAY_US"
                    loc_filter = None
                elif mkt.startswith("EBAY_US_"):
                    actual_mkt = "EBAY_US"
                    loc_filter = mkt.split("_")[-1]  # 'JP', 'TW', 'CN', 'KR' 추출
                else:
                    actual_mkt = mkt
                    loc_filter = None

                data = search_ebay_once(query, actual_mkt, item_location_country=loc_filter)
                had_success = True
            except requests.HTTPError as e:
                return jsonify({"error": str(e)}), 400

            total_matching += data.get("total", 0)
            warnings.extend(data.get("warnings", []) or [])

            for it_summary in data.get("itemSummaries", []) or []:
                # 고정가(Buy It Now) 매물은 "price" 필드를 쓰지만, 경매(auction) 매물은
                # "price"가 아예 없고 "currentBidPrice"(현재 입찰가)만 온다.
                # 예전 코드는 "price"가 없으면 그냥 continue로 건너뛰어서, eBay에 매물이
                # 있어도(total_matching > 0) 화면 목록에서는 통째로 사라지는 버그가 있었다.
                price_obj = it_summary.get("price")
                is_auction = False
                if not price_obj or price_obj.get("value") is None:
                    price_obj = it_summary.get("currentBidPrice") or {}
                    is_auction = True

                # 핵심 버그: X-EBAY-C-ENDUSERCTX(contextualLocation=KR) 헤더 때문에
                # eBay가 "미국(EBAY_US)" 매물 가격을 원래 USD가 아니라 구매자 위치(한국) 기준
                # KRW로 자동 환산해서 내려주는 경우가 있다. 이러면 currency가 "KRW"로 찍히고
                # 우리 환율표에는 KRW가 없어서 usd_equiv가 계산 안 되고, 그 결과 이 매물들이
                # 전부 통계(최저/최고/중앙값)와 목록에서 사실상 빠지는 문제가 있었다.
                # eBay는 이렇게 자동 환산할 때 원래(판매자) 통화/가격을 convertedFromValue /
                # convertedFromCurrency 필드에 같이 내려주므로, 있으면 그걸 우선 사용해서
                # 원래 통화(대부분 USD) 기준으로 정확히 계산한다.
                if price_obj.get("convertedFromCurrency") and price_obj.get("convertedFromValue") is not None:
                    val = price_obj.get("convertedFromValue")
                    cur = price_obj.get("convertedFromCurrency")
                else:
                    val = price_obj.get("value")
                    cur = price_obj.get("currency")
                if val is None:
                    # 가격 정보 자체가 없는 매물(문의 후 견적 등)만 진짜로 건너뛴다.
                    continue
                try:
                    val = float(val)
                except (TypeError, ValueError):
                    continue
                rate = fx.get(cur)
                usd_equiv = round(val * rate, 2) if rate else None
                listings.append({
                    "item_id": extract_item_id(it_summary),
                    "title": it_summary.get("title"),
                    "price": val,
                    "currency": cur,
                    "usd_equiv": usd_equiv,
                    "marketplace": mkt,
                    "marketplace_label": MARKETPLACES[mkt]["label"],
                    "condition": it_summary.get("condition"),
                    "itemWebUrl": it_summary.get("itemWebUrl"),
                    "seller": (it_summary.get("seller") or {}).get("username"),
                    "is_auction": is_auction,
                    # 월드와이드로 검색해도 실제 이 매물이 어느 나라에서 출고되는지는
                    # eBay가 itemLocation.country로 알려준다 (예: "US", "JP", "KR" 등)
                    "item_location_country": (it_summary.get("itemLocation") or {}).get("country"),
                })

            time.sleep(0.15)  # eBay rate limit 여유 (마켓플레이스별 호출 사이)

        if not had_success:
            results.append({"row_id": row_id, "query": query, "error": "모든 마켓플레이스 조회 실패: " + "; ".join(warnings)})
            continue

        raw_count = len(listings)
        listings = dedupe_listings(listings)
        duplicates_removed = raw_count - len(listings)

        # 통계(최저/최고/중앙값)는 중복 제거 후 남은, 경매 아닌 매물만 대상으로 계산
        usd_equiv_prices = [l["usd_equiv"] for l in listings if l["usd_equiv"] is not None and not l["is_auction"]]

        listings.sort(key=lambda l: l["usd_equiv"] if l["usd_equiv"] is not None else float("inf"))

        stat = None
        if usd_equiv_prices:
            median_val, was_trimmed = trimmed_median(usd_equiv_prices)
            stat = {
                "count": len(usd_equiv_prices),
                "min": min(usd_equiv_prices),
                "max": max(usd_equiv_prices),
                "median": median_val,
                "median_trimmed": was_trimmed,  # True면 최저1/최고1 제외 후 계산된 값
            }
            record_price_history(
                part_no=(it.get("part_no") or query),
                brand=it.get("brand"),
                query=query,
                stat=stat,
                marketplaces=marketplaces,
            )

        results.append({
            "row_id": row_id,
            "query": query,
            "marketplaces_searched": [MARKETPLACES[m]["label"] for m in marketplaces],
            "total_matching": total_matching,
            "duplicates_removed": duplicates_removed,
            "listings": listings[:15],
            "usd_stats": stat,
            "warnings": warnings,
            "web_search_url": f"https://www.ebay.com/sch/i.html?_nkw={quote(query)}",
        })

    return jsonify({"results": results})


@app.route("/")
def index():
    return send_from_directory(resource_path("static"), "index.html")


# ---------------------------------------------------------------------------
# 3) 가격 이력(가격추이) 조회/삭제 API — search-ebay에서 자동 기록된 데이터를
#    프론트엔드 "가격 추이" 탭에서 읽고 지우는 용도.
# ---------------------------------------------------------------------------
@app.route("/api/price-history/parts", methods=["GET"])
def list_price_history_parts():
    """이력이 하나라도 있는 부품번호 목록 (드롭다운 채우기용)."""
    conn = get_db()
    rows = conn.execute("""
        SELECT part_no,
               COUNT(*) AS cnt,
               MAX(created_at) AS last_at,
               (SELECT median_usd FROM price_history p2
                WHERE p2.part_no = p1.part_no
                ORDER BY created_at DESC LIMIT 1) AS last_median
        FROM price_history p1
        GROUP BY part_no
        ORDER BY last_at DESC
    """).fetchall()
    conn.close()
    return jsonify({
        "parts": [
            {"part_no": r["part_no"], "count": r["cnt"], "last_at": r["last_at"], "last_median": r["last_median"]}
            for r in rows
        ]
    })


@app.route("/api/price-history", methods=["GET"])
def get_price_history():
    """특정 부품번호의 전체 이력 (오래된 순)."""
    part_no = (request.args.get("part_no") or "").strip().upper()
    if not part_no:
        return jsonify({"error": "part_no 파라미터가 필요합니다"}), 400
    conn = get_db()
    rows = conn.execute(
        "SELECT id, part_no, brand, query, count, min_usd, max_usd, median_usd, marketplaces, created_at "
        "FROM price_history WHERE part_no = ? ORDER BY created_at ASC",
        (part_no,),
    ).fetchall()
    conn.close()
    return jsonify({"part_no": part_no, "history": [dict(r) for r in rows]})


@app.route("/api/price-history/<int:entry_id>", methods=["DELETE"])
def delete_price_history_entry(entry_id):
    """이력 한 줄 삭제 (가격 추이 테이블의 × 버튼용)."""
    conn = get_db()
    conn.execute("DELETE FROM price_history WHERE id = ?", (entry_id,))
    conn.commit()
    conn.close()
    return jsonify({"deleted": entry_id})


if __name__ == "__main__":
    port = int(os.getenv("PORT", "5050"))
    host = os.getenv("HOST", "127.0.0.1")  # 클라우드 배포 시 Render가 자동으로 0.0.0.0 지정
    print(f"eBay 가격 조회 도구 실행 중: http://{host}:{port}  (환경: {EBAY_ENV})")
    if not EBAY_CLIENT_ID or not EBAY_CLIENT_SECRET:
        print(f"경고: {os.path.join(BASE_DIR, '.env')} 에 EBAY_CLIENT_ID/EBAY_CLIENT_SECRET이 없습니다.")

    # exe로 실행했을 때 자동으로 기본 브라우저를 열어준다 (로컬 python 실행 시에도 동일하게 동작)
    if os.getenv("NO_AUTO_OPEN") != "1":
        threading.Timer(1.0, lambda: webbrowser.open(f"http://{host}:{port}")).start()

    app.run(host=host, port=port, debug=False)
